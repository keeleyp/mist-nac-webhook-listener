#!/usr/bin/env python3
"""
Juniper Mist NAC Accounting Webhook Listener
Listens on port 8001 for nac-accounting webhook events and logs to dated Excel files.
"""

import json
import os
import logging
from datetime import datetime, date, timezone
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import Workbook, load_workbook

# ── Configuration ────────────────────────────────────────────────────────────
PORT = 8001
DEBUG = True
OUTPUT_DIR = "."  # Directory to store Excel files

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG if DEBUG else logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    "Received At",
    "Type",
    "Username",
    "MAC",
    "Client IP",
    "Client Type",
    "SSID",
    "AP",
    "BSSID",
    "NAS IP",
    "NAS Vendor",
    "Session ID",
    "Session Unique ID",
    "Multi Session ID",
    "Site ID",
    "Org ID",
    "Device Cert Expiry",
    "Timestamp (ms)",
    "Session Duration (mins)",
    "RX Bytes",
    "TX Bytes",
    "RX Pkts",
    "TX Pkts",
    "Terminate Cause",
]

# ── Excel helpers ─────────────────────────────────────────────────────────────

def get_excel_path(for_date: date = None) -> str:
    d = for_date or date.today()
    return os.path.join(OUTPUT_DIR, f"mist_nac_{d.strftime('%Y-%m-%d')}.xlsx")


def get_or_create_workbook(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "NAC Accounting"
        ws.append(COLUMNS)
        # Basic column width hints
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
                len(col_name) + 2, 14
            )
        log.info("Created new workbook: %s", path)
    return wb, ws


def append_event(event: dict) -> None:
    today_path = get_excel_path()
    wb, ws = get_or_create_workbook(today_path)

    event_type = event.get("type", "")
    received_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ts_ms = event.get("timestamp")
    ts_human = (
        datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        if ts_ms
        else ""
    )

    row = [
        received_at,
        event_type,
        event.get("username", ""),
        event.get("mac", ""),
        event.get("client_ip", ""),
        event.get("client_type", ""),
        event.get("ssid", ""),
        event.get("ap", ""),
        event.get("bssid", ""),
        event.get("nas_ip", ""),
        event.get("nas_vendor", ""),
        event.get("session_id", ""),
        event.get("session_unique_id", ""),
        event.get("multi_session_id", ""),
        event.get("site_id", ""),
        event.get("org_id", ""),
        event.get("device_cert_expiry", ""),
        ts_human,
        event.get("session_duration_in_mins", ""),
        event.get("rx_bytes", ""),
        event.get("tx_bytes", ""),
        event.get("rx_pkts", ""),
        event.get("tx_pkts", ""),
        event.get("terminate_cause", ""),
    ]

    ws.append(row)
    wb.save(today_path)
    log.debug("Saved row to %s", today_path)

# ── Debug output ──────────────────────────────────────────────────────────────

def print_event_summary(event: dict) -> None:
    if not DEBUG:
        return
    event_type = event.get("type", "UNKNOWN")
    divider = "─" * 60
    print(divider)
    print(f"  Event Type : {event_type}")
    print(f"  Username   : {event.get('username', 'N/A')}")
    print(f"  MAC        : {event.get('mac', 'N/A')}")
    print(f"  Client IP  : {event.get('client_ip', 'N/A')}")
    print(f"  SSID       : {event.get('ssid', 'N/A')}")
    print(f"  AP         : {event.get('ap', 'N/A')}")
    print(f"  Session ID : {event.get('session_id', 'N/A')}")
    print(f"  NAS IP     : {event.get('nas_ip', 'N/A')}")
    ts_ms = event.get("timestamp")
    if ts_ms:
        ts_str = datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        print(f"  Timestamp  : {ts_str}")
    if event_type == "NAC_ACCOUNTING_STOP":
        print(f"  Duration   : {event.get('session_duration_in_mins', 'N/A')} min(s)")
        print(f"  RX Bytes   : {event.get('rx_bytes', 'N/A')}")
        print(f"  TX Bytes   : {event.get('tx_bytes', 'N/A')}")
        print(f"  Term Cause : {event.get('terminate_cause', 'N/A')}")
    print(divider)

# ── HTTP handler ──────────────────────────────────────────────────────────────

class WebhookHandler(BaseHTTPRequestHandler):

    def do_POST(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length)

        try:
            payload = json.loads(body)
        except json.JSONDecodeError as exc:
            log.warning("Failed to parse JSON body: %s", exc)
            self._respond(400, "Invalid JSON")
            return

        topic = payload.get("topic", "")
        if topic != "nac-accounting":
            log.debug("Ignoring non-NAC topic: %s", topic)
            self._respond(200, "OK (ignored)")
            return

        events = payload.get("events", [])
        log.info("Received %d event(s) for topic '%s'", len(events), topic)

        for event in events:
            print_event_summary(event)
            append_event(event)

        self._respond(200, "OK")

    def _respond(self, code: int, message: str) -> None:
        self.send_response(code)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write(message.encode())

    def log_message(self, fmt, *args):
        # Route HTTP access log through Python logger
        log.debug("HTTP %s", fmt % args)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Mist NAC Accounting Webhook Listener")
    print(f"  Port      : {PORT}")
    print(f"  Debug     : {DEBUG}")
    print(f"  Output    : {os.path.abspath(OUTPUT_DIR)}/mist_nac_YYYY-MM-DD.xlsx")
    print(f"  Started   : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    server = HTTPServer(("0.0.0.0", PORT), WebhookHandler)
    log.info("Listening on 0.0.0.0:%d …", PORT)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        log.info("Shutting down.")
        server.server_close()


if __name__ == "__main__":
    main()
